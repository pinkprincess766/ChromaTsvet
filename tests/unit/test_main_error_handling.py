import csv
import json
from pathlib import Path
import sqlite3
from tempfile import TemporaryDirectory
from types import SimpleNamespace
import unittest
from unittest.mock import patch

import numpy as np
from openpyxl import load_workbook
from PyQt5.QtCore import QSettings
from PyQt5.QtGui import QKeySequence
from PyQt5.QtWidgets import QApplication, QMessageBox

import python_analyzer.main as main
from python_analyzer.analysis.models import ReferencePeak
from python_analyzer.analysis.peak_review import (
    PEAK_REVIEW_ACCEPTED,
    PEAK_REVIEW_REJECTED,
    PeakReview,
)
from python_analyzer.core.identification import (
    MatchResult,
    SpectrumIdentifier,
    compute_peak_based_score,
    find_peak_matches,
    normalize_data_type,
    peak_to_reference_peak,
)
from python_analyzer.gui.recent_files import (
    load_last_directory,
    load_recent_files,
    remember_recent_file,
)
from tests.support.gui_fakes import (
    FakeGraphExporter,
    FakeIdentifier,
    create_test_plot_image,
)


class MainWindowErrorHandlingTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QApplication.instance() or QApplication([])
        QSettings.setDefaultFormat(QSettings.IniFormat)
        QSettings.setPath(
            QSettings.IniFormat,
            QSettings.UserScope,
            "/tmp/chromatsvet-error-handling-tests",
        )

    def setUp(self):
        main.app_settings().clear()
        self.identifier_patch = patch.object(main, "SpectrumIdentifier", FakeIdentifier)
        self.process_patch = patch.object(
            main.spectrometer_rust,
            "process_signal",
            return_value={"spectrum": [1.0, 2.0], "peaks": []},
        )
        self.identifier_patch.start()
        self.process_signal = self.process_patch.start()
        self.window = main.MainWindow()
        self.assertIsNone(self.window.current_data)
        self.process_signal.assert_not_called()
        self.window.current_data = [0.1, 0.3, 0.8, 0.2]

    def tearDown(self):
        self.window.close()
        self.process_patch.stop()
        self.identifier_patch.stop()

    def create_test_plot_image(self, directory):
        return create_test_plot_image(directory)

    def test_log_levels_are_prefixed_and_colored(self):
        self.window.log("Warning entry", level="warning")
        self.window.log("Error entry", level="error")

        self.assertIn("[WARN] Warning entry", self.window.log_history[-2])
        self.assertIn("[ERROR] Error entry", self.window.log_history[-1])
        self.assertIn("Application started", self.window.embedded_log_view.toPlainText())

        embedded_html = self.window.embedded_log_view.toHtml().lower()
        self.assertIn("#f5a623", embedded_html)
        self.assertIn("#ff5c5c", embedded_html)

        log_window = main.LogWindow(self.window)
        html = log_window.log_view.toHtml().lower()
        self.assertIn("#f5a623", html)
        self.assertIn("#ff5c5c", html)
        log_window.close()

    def test_embedded_log_copy_and_clear_controls(self):
        self.window.log("Copy this entry")

        self.window.log_copy_button.click()
        self.assertEqual(
            QApplication.clipboard().text(),
            self.window.embedded_log_view.toPlainText(),
        )

        self.window.log_clear_button.click()
        self.assertEqual(self.window.log_history, [])
        self.assertEqual(self.window.embedded_log_view.toPlainText(), "")

    def test_analysis_failure_is_reported_without_raising(self):
        self.process_signal.side_effect = RuntimeError("Rust failure")

        with patch.object(QMessageBox, "critical", return_value=QMessageBox.Ok) as critical:
            self.window.run_analysis()

        critical.assert_called_once()
        self.assertIn("[ERROR]", self.window.log_history[-1])
        self.assertIn("RuntimeError: Rust failure", self.window.log_history[-1])
        self.assertEqual(self.window.status_bar.currentMessage(), "Spectrum analysis failed")

    def test_processing_warnings_are_reported_in_analysis_console(self):
        self.process_signal.return_value = {
            "spectrum": [0.0, 1.0],
            "peaks": [],
            "processing_warnings": [
                "area_normalization_skipped",
                "/home/scientist/private/raw.csv",
            ],
        }

        self.window.run_analysis()

        log_text = "\n".join(self.window.log_history)
        self.assertIn("Processing warning", log_text)
        self.assertIn("area normalization was skipped", log_text)
        self.assertIn("Rust reported '.../raw.csv'", log_text)
        self.assertNotIn("/home/scientist/private", log_text)

    def test_peak_review_warnings_are_reported_in_analysis_console(self):
        weak_peak = SimpleNamespace(
            position=1.0,
            frequency=250.0,
            intensity=1.0,
            prominence=0.1,
            baseline_level=0.0,
            left_base=0.0,
            right_base=0.0,
            width=0.0,
            width_hz=0.0,
            area=0.0,
            noise=0.1,
            snr=1.0,
            is_global_max=False,
        )
        self.process_signal.return_value = {
            "spectrum": [0.0, 1.0],
            "peaks": [weak_peak],
        }

        self.window.run_analysis()

        log_text = "\n".join(self.window.log_history)
        self.assertIn("Peak review", log_text)
        self.assertIn("1 peak requires attention", log_text)
        self.assertIn("low SNR", log_text)
        self.assertIn("unknown width", log_text)

    def test_file_read_failure_is_reported_without_raising(self):
        missing_file = "/tmp/chromatsvet-file-that-does-not-exist.csv"

        with (
            patch.object(main.QFileDialog, "getOpenFileName", return_value=(missing_file, "")),
            patch.object(QMessageBox, "warning", return_value=QMessageBox.Ok) as warning,
        ):
            self.window.load_file()

        warning.assert_called_once()
        self.assertEqual(warning.call_args.args[1], "File not found")
        self.assertIn("moved, renamed, or deleted", warning.call_args.args[2])
        self.assertNotIn(missing_file, warning.call_args.args[2])
        self.assertIn("[ERROR]", self.window.log_history[-1])
        self.assertIn("FileNotFoundError", self.window.log_history[-1])
        self.assertNotIn(missing_file, self.window.log_history[-1])

    def test_decode_failure_message_is_actionable(self):
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "bad_encoding.csv"
            file_path.write_bytes(b"intensity\n\xff\n")

            with (
                patch.object(
                    main.QFileDialog,
                    "getOpenFileName",
                    return_value=(str(file_path), ""),
                ),
                patch.object(
                    QMessageBox, "warning", return_value=QMessageBox.Ok
                ) as warning,
            ):
                self.window.load_file()

        warning.assert_called_once()
        self.assertEqual(warning.call_args.args[1], "Unsupported text encoding")
        self.assertIn("UTF-8 CSV/TXT", warning.call_args.args[2])
        self.assertNotIn(temp_dir, warning.call_args.args[2])
        self.assertIn("UnicodeDecodeError", self.window.log_history[-1])
        self.assertNotIn(temp_dir, self.window.log_history[-1])

    def test_invalid_substance_input_is_reported(self):
        answers = [("Test", True), ("T", True), ("1,,2", True)]

        with (
            patch.object(main.QInputDialog, "getText", side_effect=answers),
            patch.object(QMessageBox, "warning", return_value=QMessageBox.Ok) as warning,
        ):
            self.window.add_substance()

        warning.assert_called_once()
        self.assertIn("[ERROR]", self.window.log_history[-1])
        self.assertIn("intensity values cannot be empty", self.window.log_history[-1])

    def test_load_file_shows_format_warning_for_invalid_table(self):
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "broken.csv"
            file_path.write_text(
                "position;intensity\n0;1.0\n1;2.0;unexpected\n",
                encoding="utf-8",
            )

            with (
                patch.object(
                    main.QFileDialog,
                    "getOpenFileName",
                    return_value=(str(file_path), ""),
                ),
                patch.object(
                    QMessageBox, "warning", return_value=QMessageBox.Ok
                ) as warning,
            ):
                self.window.load_file()

        warning.assert_called_once()
        self.assertIn("How to fix", warning.call_args.args[2])
        self.assertIn("Invalid spectrum file format", self.window.log_history[-1])

    def test_load_file_remembers_recent_file_directory_and_status(self):
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "sample.csv"
            file_path.write_text("intensity\n0.1\n0.3\n0.8\n0.2\n", encoding="utf-8")

            with (
                patch.object(
                    main.QFileDialog,
                    "getOpenFileName",
                    return_value=(str(file_path), "CSV (*.csv)"),
                ),
                patch.object(QMessageBox, "warning", return_value=QMessageBox.Ok),
            ):
                self.window.load_file()

            self.assertEqual(load_recent_files(self.window.settings), [str(file_path)])
            self.assertEqual(load_last_directory(self.window.settings), temp_dir)
            self.assertEqual(self.window.status_source_label.text(), "sample.csv")
            self.assertIn("4 points", self.window.status_details_label.text())
            self.assertIn("0 peaks", self.window.status_details_label.text())
            self.assertIn("analyzed", self.window.status_details_label.text())
            self.assertIn("threshold=0.050", self.window.status_details_label.text())

            with patch.object(
                main.QFileDialog,
                "getOpenFileName",
                return_value=("", ""),
            ) as open_dialog:
                self.window.load_file()

            self.assertEqual(open_dialog.call_args.args[2], temp_dir)

    def test_overlay_file_is_analyzed_and_can_be_cleared(self):
        primary_result = {
            "spectrum": [1.0, 2.0],
            "frequency_axis": [0.0, 250.0],
            "sample_rate": 1000.0,
            "peaks": [],
        }
        overlay_result = {
            "spectrum": [0.5, 1.5],
            "frequency_axis": [0.0, 250.0],
            "sample_rate": 1000.0,
            "peaks": [],
        }
        self.process_signal.side_effect = [primary_result, overlay_result]
        self.window.run_analysis()

        with TemporaryDirectory() as temp_dir:
            overlay_path = Path(temp_dir) / "overlay.csv"
            overlay_path.write_text("intensity\n0.2\n0.6\n0.4\n0.1\n", encoding="utf-8")

            with patch.object(
                main.QFileDialog,
                "getOpenFileName",
                return_value=(str(overlay_path), "CSV (*.csv)"),
            ):
                self.window.load_overlay_file()

        self.assertEqual(self.window.overlay_file_name, "overlay.csv")
        self.assertEqual(self.window.overlay_spectrum_values.tolist(), [0.5, 1.5])
        self.assertIn("overlay: overlay.csv", self.window.status_details_label.text())
        self.assertTrue(self.window.clear_overlay_action.isEnabled())

        self.window.clear_overlay_spectrum()

        self.assertIsNone(self.window.overlay_data)
        self.assertNotIn("overlay:", self.window.status_details_label.text())

    def test_missing_recent_file_is_removed_without_path_leak(self):
        with TemporaryDirectory() as temp_dir:
            missing_file = Path(temp_dir) / "missing.csv"
            self.window.recent_files = remember_recent_file(
                self.window.settings,
                missing_file,
            )
            self.window._refresh_recent_files_menu()

            with patch.object(
                QMessageBox,
                "warning",
                return_value=QMessageBox.Ok,
            ) as warning:
                self.window.open_recent_file(str(missing_file))

            warning.assert_called_once()
            self.assertEqual(load_recent_files(self.window.settings), [])
            self.assertIn("Recent file unavailable", self.window.log_history[-1])
            self.assertIn("missing.csv", self.window.log_history[-1])
            self.assertNotIn(temp_dir, self.window.log_history[-1])

    def test_clear_recent_files_removes_menu_entries(self):
        with TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "sample.csv"
            self.window.recent_files = remember_recent_file(
                self.window.settings,
                file_path,
            )
            self.window._refresh_recent_files_menu()

        self.assertTrue(load_recent_files(self.window.settings))

        self.window.clear_recent_files()

        self.assertEqual(load_recent_files(self.window.settings), [])
        self.assertEqual(self.window.recent_files_menu.actions()[0].text(), "No recent files")

    def test_workflow_shortcuts_are_registered(self):
        def shortcut_texts(action):
            return [
                shortcut.toString(QKeySequence.PortableText)
                for shortcut in action.shortcuts()
            ]

        self.assertEqual(shortcut_texts(self.window.open_file_action), ["Ctrl+O"])
        self.assertEqual(
            shortcut_texts(self.window.run_analysis_action),
            ["Ctrl+R", "F5"],
        )
        self.assertEqual(shortcut_texts(self.window.export_pdf_action), ["Ctrl+E"])
        self.assertEqual(
            shortcut_texts(self.window.export_peaks_action),
            ["Ctrl+Shift+E"],
        )
        self.assertEqual(
            shortcut_texts(self.window.analysis_settings_action),
            ["Ctrl+,"],
        )
        self.assertIn("Ctrl+O", self.window.btn_open.toolTip())
        self.assertIn("F5", self.window.btn_run.toolTip())

    def test_database_failure_is_reported(self):
        self.window.identifier.clear_database = lambda: False

        with (
            patch.object(QMessageBox, "question", return_value=QMessageBox.Yes),
            patch.object(QMessageBox, "critical", return_value=QMessageBox.Ok) as critical,
        ):
            self.window.clear_database()

        critical.assert_called_once()
        self.assertIn("[ERROR]", self.window.log_history[-1])
        self.assertIn("RuntimeError", self.window.log_history[-1])

    def test_pdf_permission_failure_is_reported(self):
        self.window.run_analysis()

        with (
            patch.object(
                main.QFileDialog,
                "getSaveFileName",
                return_value=("/restricted/report.pdf", ""),
            ),
            patch.object(
                main.tempfile,
                "NamedTemporaryFile",
                side_effect=PermissionError("access denied"),
            ),
            patch.object(QMessageBox, "warning", return_value=QMessageBox.Ok) as warning,
        ):
            self.window.export_pdf()

        warning.assert_called_once()
        self.assertIn("[ERROR]", self.window.log_history[-1])
        self.assertIn("PermissionError: access denied", self.window.log_history[-1])

    def test_pdf_report_writes_file(self):
        self.window.run_analysis()

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "report.pdf"
            plot_path = self.create_test_plot_image(temp_dir)
            with (
                patch.object(
                    main.QFileDialog,
                    "getSaveFileName",
                    return_value=(str(output_path), "PDF (*.pdf)"),
                ),
                patch.object(
                    self.window,
                    "_render_plot_snapshot",
                    return_value=str(plot_path),
                ),
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
            ):
                self.window.export_pdf()

            self.assertTrue(output_path.is_file())
            self.assertGreater(output_path.stat().st_size, 0)

        self.assertIn("PDF report created:", self.window.log_history[-1])

    def test_html_report_writes_file(self):
        self.window.run_analysis()

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "report.html"
            plot_path = self.create_test_plot_image(temp_dir)
            with (
                patch.object(
                    main.QFileDialog,
                    "getSaveFileName",
                    return_value=(str(output_path), "HTML (*.html)"),
                ),
                patch.object(
                    self.window,
                    "_render_plot_snapshot",
                    return_value=str(plot_path),
                ),
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
            ):
                self.window.export_html()

            self.assertTrue(output_path.is_file())
            html = output_path.read_text(encoding="utf-8")

        self.assertIn("ChromaTsvet Analysis Report", html)
        self.assertIn("In-memory data", html)
        self.assertIn("data:image/png;base64,", html)
        self.assertIn("HTML report created:", self.window.log_history[-1])

    def test_excel_report_writes_workbook(self):
        self.window.run_analysis()

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "report.xlsx"
            with (
                patch.object(
                    main.QFileDialog,
                    "getSaveFileName",
                    return_value=(str(output_path), "Excel Workbook (*.xlsx)"),
                ),
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
            ):
                self.window.export_excel()

            workbook = load_workbook(output_path, data_only=True)

        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "Parameters", "Processing Passport", "Peaks", "Matches"],
        )
        self.assertEqual(workbook["Summary"]["A5"].value, "Source file")
        self.assertEqual(workbook["Summary"]["B5"].value, "In-memory data")
        self.assertEqual(workbook["Processing Passport"]["A1"].value, "Field")
        self.assertEqual(workbook["Processing Passport"]["A4"].value, "Rust core")
        self.assertIn("Excel workbook created:", self.window.log_history[-1])

    def test_graph_png_export_writes_image_and_remembers_directory(self):
        self.window.run_analysis()

        with TemporaryDirectory() as temp_dir:
            output_path_without_suffix = Path(temp_dir) / "spectrum_graph"
            expected_output_path = output_path_without_suffix.with_suffix(".png")
            with (
                patch.object(
                    main.QFileDialog,
                    "getSaveFileName",
                    return_value=(str(output_path_without_suffix), "PNG Image (*.png)"),
                ),
                patch(
                    "python_analyzer.gui.main_window.pg_exporters.ImageExporter",
                    FakeGraphExporter,
                ),
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
            ):
                self.window.export_graph_png()

            self.assertTrue(expected_output_path.is_file())
            self.assertEqual(expected_output_path.read_bytes(), FakeGraphExporter.payload)
            self.assertEqual(load_last_directory(self.window.settings), temp_dir)
            self.assertIn("Graph PNG exported: spectrum_graph.png", self.window.log_history[-1])
            self.assertNotIn(temp_dir, self.window.log_history[-1])

    def test_graph_svg_export_writes_image_and_updates_log(self):
        self.window.run_analysis()

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "spectrum_graph.svg"
            with (
                patch.object(
                    main.QFileDialog,
                    "getSaveFileName",
                    return_value=(str(output_path), "SVG Image (*.svg)"),
                ),
                patch(
                    "python_analyzer.gui.main_window.pg_exporters.SVGExporter",
                    FakeGraphExporter,
                ),
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
            ):
                self.window.export_graph_svg()

            self.assertTrue(output_path.is_file())
            self.assertEqual(output_path.read_bytes(), FakeGraphExporter.payload)
            self.assertEqual(load_last_directory(self.window.settings), temp_dir)
            self.assertIn("Graph SVG exported: spectrum_graph.svg", self.window.log_history[-1])
            self.assertNotIn(temp_dir, self.window.log_history[-1])

    def test_graph_export_is_disabled_without_analysis_results(self):
        self.assertIsNone(self.window.current_result)

        with (
            patch.object(main.QFileDialog, "getSaveFileName") as save_dialog,
            patch.object(QMessageBox, "warning", return_value=QMessageBox.Ok) as warning,
        ):
            self.window.export_graph_png()

        warning.assert_called_once()
        save_dialog.assert_not_called()
        self.assertIn("[WARN]", self.window.log_history[-1])

    def test_peak_export_is_disabled_without_detected_peaks(self):
        self.assertFalse(self.window.btn_export_peaks.isEnabled())
        self.assertFalse(self.window.export_peaks_action.isEnabled())

        with (
            patch.object(main.QFileDialog, "getSaveFileName") as save_dialog,
            patch.object(QMessageBox, "warning", return_value=QMessageBox.Ok) as warning,
        ):
            self.window.export_peaks_csv()

        warning.assert_called_once()
        save_dialog.assert_not_called()
        self.assertIn("[WARN]", self.window.log_history[-1])

    def test_peak_export_writes_csv_and_updates_log(self):
        peak = SimpleNamespace(
            frequency=62.0,
            position=124.0,
            intensity=0.854,
            width=12.3,
            width_hz=6.15,
            area=4.21,
            snr=18.7,
        )
        self.process_signal.return_value = {"spectrum": [0.0, 1.0], "peaks": [peak]}
        self.window.run_analysis()
        self.assertTrue(self.window.btn_export_peaks.isEnabled())
        self.assertTrue(self.window.export_peaks_action.isEnabled())

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "peaks.csv"
            with (
                patch.object(
                    main.QFileDialog,
                    "getSaveFileName",
                    return_value=(str(output_path), "CSV (*.csv)"),
                ),
                patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
            ):
                self.window.export_peaks_csv()

            with output_path.open(newline="", encoding="utf-8") as csv_file:
                rows = list(csv.reader(csv_file))

        self.assertEqual(
            rows[0],
            [
                "source_file",
                "sample_rate_hz",
                "filter_type",
                "baseline",
                "normalization",
                "spectrum_smoothing",
                "spectrum_smoothing_method",
                "spectrum_smoothing_window",
                "peak_min_snr",
                "data_type",
                "peak_frequency_tolerance_hz",
                "frequency_hz",
                "position_bin",
                "intensity",
                "width_bins",
                "width_hz",
                "area",
                "snr",
                "review_status",
                "review_reason",
            ],
        )
        self.assertEqual(
            rows[1],
            [
                "In-memory data",
                "1000.0",
                "median",
                "improved",
                "none",
                "False",
                "none",
                "0",
                "0.0",
                "generic",
                "5.0",
                "62.0",
                "124.0",
                "0.854",
                "12.3",
                "6.15",
                "4.21",
                "18.7",
                "accepted",
                "accepted",
            ],
        )
        self.assertIn("Peak list exported:", self.window.log_history[-1])

    def test_detected_peaks_are_shown_in_table_and_plot(self):
        peaks = [
            SimpleNamespace(
                frequency=6.0,
                position=12.0,
                intensity=0.42,
                width=2.5,
                width_hz=1.25,
                area=1.2,
                snr=8.0,
            ),
            SimpleNamespace(
                frequency=12.25,
                position=24.5,
                intensity=0.91,
                width=3.5,
                width_hz=1.75,
                area=2.4,
                snr=15.0,
            ),
        ]
        self.process_signal.return_value = {
            "spectrum": [0.0, 0.42, 0.1, 0.91],
            "peaks": peaks,
        }

        self.window.run_analysis()

        self.assertEqual(self.window.peak_table.rowCount(), 2)
        self.assertEqual(self.window.peak_table.item(0, 0).text(), "6")
        self.assertEqual(self.window.peak_table.item(0, 1).text(), "12")
        self.assertEqual(self.window.peak_table.item(0, 2).text(), "0.42")
        self.assertEqual(self.window.peak_table.item(0, 4).text(), "1.25")
        self.assertEqual(self.window.peak_table.item(0, 8).text(), "accepted")
        self.assertEqual(self.window.peak_table.item(1, 0).text(), "12.25")
        self.assertEqual(self.window.results_tabs.tabText(0), "Detected Peaks (2)")
        self.assertGreaterEqual(
            len(self.window.plot.getPlotItem().listDataItems()),
            2,
        )

    def test_method_preset_can_be_saved_and_applied(self):
        self.window.set_analysis_settings(
            baseline_enabled=True,
            baseline_method="improved",
            peak_threshold=0.02,
            peak_prominence=0.4,
            peak_distance=6,
            peak_min_snr=12.0,
            filter_type="none",
            filter_params={},
            sample_rate=2500.0,
            window_type="hamming",
            normalize_area=True,
            spectrum_smoothing_enabled=True,
            spectrum_smoothing_method="median",
            spectrum_smoothing_window=11,
            peak_frequency_tolerance=2.0,
            data_type="raman",
        )

        saved_name = self.window.save_current_method_preset(" Raman QC ")
        self.window.set_analysis_settings(
            baseline_enabled=False,
            baseline_method="simple",
            peak_threshold=0.5,
            peak_prominence=0.0,
            peak_distance=1,
            filter_type="median",
            filter_params={"window_size": 5},
            sample_rate=1000.0,
            window_type="rectangular",
            normalize_area=False,
            peak_frequency_tolerance=10.0,
            data_type="generic",
        )

        self.assertEqual(saved_name, "Raman QC")
        self.assertIn("Raman QC", self.window.list_method_presets())
        self.assertTrue(self.window.apply_method_preset("Raman QC"))
        self.assertEqual(self.window.sample_rate, 2500.0)
        self.assertEqual(self.window.window_type, "hamming")
        self.assertEqual(self.window.data_type, "raman")
        self.assertEqual(self.window.current_method_preset_name, "Raman QC")

    def test_rejected_peak_is_not_stored_in_reference_library(self):
        accepted_peak = SimpleNamespace(
            frequency=10.0,
            intensity=1.0,
            width=2.0,
            width_hz=1.0,
            area=3.0,
            snr=20.0,
        )
        rejected_peak = SimpleNamespace(
            frequency=20.0,
            intensity=0.5,
            width=2.0,
            width_hz=1.0,
            area=1.0,
            snr=4.0,
        )
        self.window.current_peaks = [accepted_peak, rejected_peak]
        self.window.current_peak_reviews = [
            PeakReview(PEAK_REVIEW_ACCEPTED, "accepted"),
            PeakReview(PEAK_REVIEW_REJECTED, "rejected by user", user_modified=True),
        ]

        with (
            patch.object(
                main.QInputDialog,
                "getText",
                side_effect=[
                    ("Water", True),
                    ("H2O", True),
                    ("1.0, 2.0", True),
                ],
            ),
            patch.object(
                self.window.identifier,
                "add_reference",
                return_value=True,
            ) as add_reference,
            patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
        ):
            self.window.add_substance()

        add_reference.assert_called_once()
        self.assertEqual(len(add_reference.call_args.kwargs["peaks"]), 1)
        self.assertEqual(add_reference.call_args.kwargs["peaks"][0].frequency, 10.0)

    def test_plot_mouse_zoom_is_enabled(self):
        view_box = self.window.plot.getPlotItem().getViewBox()

        self.assertEqual(view_box.state["mouseEnabled"], [True, True])
        self.assertEqual(view_box.state["mouseMode"], main.pg.ViewBox.RectMode)
        self.assertIn("zoom", self.window.plot.toolTip().lower())

    def test_toolbar_logo_is_loaded(self):
        self.assertTrue(main.APP_LOGO_PATH.is_file())
        self.assertFalse(self.window.logo_label.isHidden())
        self.assertFalse(self.window.logo_label.pixmap().isNull())

    def test_area_normalization_is_disabled_by_default(self):
        self.assertFalse(self.window.normalize_area)
        self.window.run_analysis()
        self.assertNotIn("normalize", self.process_signal.call_args.kwargs)
        self.assertEqual(
            self.process_signal.call_args.kwargs["sample_rate"],
            main.DEFAULT_SAMPLE_RATE,
        )
        self.assertEqual(
            self.process_signal.call_args.kwargs["window_type"],
            main.DEFAULT_WINDOW_TYPE,
        )

    def test_analysis_settings_are_saved_and_passed_to_rust(self):
        self.window.set_analysis_settings(
            baseline_enabled=self.window.baseline_enabled,
            baseline_method=self.window.baseline_method,
            peak_threshold=self.window.peak_threshold,
            peak_prominence=self.window.peak_prominence,
            peak_distance=self.window.peak_distance,
            filter_type=self.window.filter_type,
            filter_params=self.window.filter_params,
            sample_rate=2_500.0,
            window_type="hamming",
            normalize_area=True,
            peak_min_snr=4.5,
            spectrum_smoothing_enabled=True,
            spectrum_smoothing_method="savgol",
            spectrum_smoothing_window=8,
            peak_frequency_tolerance=12.5,
            data_type="Raman!",
        )

        self.assertTrue(self.window.normalize_area)
        self.assertEqual(self.window.sample_rate, 2_500.0)
        self.assertEqual(self.window.window_type, "hamming")
        self.assertEqual(self.window.peak_min_snr, 4.5)
        self.assertTrue(self.window.spectrum_smoothing_enabled)
        self.assertEqual(self.window.spectrum_smoothing_method, "savgol")
        self.assertEqual(self.window.spectrum_smoothing_window, 9)
        self.assertEqual(self.window.peak_frequency_tolerance, 12.5)
        self.assertEqual(self.window.data_type, "raman")
        self.assertTrue(
            main.saved_bool(
                self.window.settings,
                "analysis/normalize_area",
                False,
            )
        )
        self.assertEqual(
            main.saved_float(
                self.window.settings,
                "analysis/sample_rate",
                0.0,
                0.001,
                10_000_000.0,
            ),
            2_500.0,
        )
        self.assertEqual(self.process_signal.call_args.kwargs["sample_rate"], 2_500.0)
        self.assertEqual(self.process_signal.call_args.kwargs["window_type"], "hamming")
        self.assertTrue(self.process_signal.call_args.kwargs["normalize"])
        self.assertEqual(self.process_signal.call_args.kwargs["min_snr"], 4.5)
        self.assertTrue(self.process_signal.call_args.kwargs["spectrum_smoothing"])
        self.assertEqual(
            self.process_signal.call_args.kwargs["spectrum_smoothing_method"],
            "savgol",
        )
        self.assertEqual(
            self.process_signal.call_args.kwargs["spectrum_smoothing_window"],
            9,
        )

    def test_analysis_dialog_applies_identification_settings(self):
        dialog = main.AnalysisSettingsDialog(self.window)
        try:
            dialog.window_type_combo.setCurrentIndex(
                dialog.window_type_combo.findData("rectangular")
            )
            dialog.spectrum_smoothing_checkbox.setChecked(True)
            dialog.spectrum_smoothing_window_spin.setValue(10)
            dialog.min_snr_spin.setValue(3.25)
            dialog.peak_tolerance_spin.setValue(17.25)
            dialog.data_type_combo.setCurrentIndex(
                dialog.data_type_combo.findData("raman")
            )

            dialog.apply_settings()
        finally:
            dialog.close()

        self.assertTrue(self.window.spectrum_smoothing_enabled)
        self.assertEqual(self.window.window_type, "rectangular")
        self.assertEqual(self.window.spectrum_smoothing_window, 11)
        self.assertEqual(self.window.peak_min_snr, 3.25)
        self.assertEqual(self.window.peak_frequency_tolerance, 17.25)
        self.assertEqual(self.window.data_type, "raman")

    def test_identification_table_uses_compared_points(self):
        self.window.identifier.find_matches = lambda spectrum: [
            SimpleNamespace(
                substance_name="Test",
                formula="T",
                score=0.75,
                compared_points=2,
            )
        ]

        self.window.run_analysis()

        self.assertEqual(
            self.window.table.horizontalHeaderItem(3).text(), "Matched"
        )
        self.assertEqual(self.window.table.item(0, 3).text(), "2")

    def test_identification_table_accepts_legacy_matched_points_result(self):
        self.window.identifier.find_matches = lambda spectrum: [
            SimpleNamespace(
                substance_name="Legacy",
                formula="L",
                score=0.5,
                matched_points=3,
            )
        ]

        self.window.run_analysis()

        self.assertEqual(self.window.table.item(0, 3).text(), "3")

    def test_add_substance_stores_object_peaks(self):
        class RecordingIdentifier(FakeIdentifier):
            def __init__(self):
                self.add_calls = []

            def add_reference(
                self,
                name,
                intensities,
                formula,
                peaks=None,
                data_type="generic",
            ):
                self.add_calls.append((name, intensities, formula, peaks, data_type))
                return True

        identifier = RecordingIdentifier()
        self.window.identifier = identifier
        self.window.current_peaks = [
            SimpleNamespace(
                frequency=100.0,
                intensity=1.5,
                width=2.0,
                width_hz=20.0,
                area=3.0,
                snr=12.0,
            )
        ]
        self.window.analysis_settings.data_type = "raman"

        answers = [("Reference", True), ("R", True), ("1.0,2.0", True)]
        with (
            patch.object(main.QInputDialog, "getText", side_effect=answers),
            patch.object(QMessageBox, "information", return_value=QMessageBox.Ok),
        ):
            self.window.add_substance()

        self.assertEqual(len(identifier.add_calls), 1)
        name, intensities, formula, peaks, data_type = identifier.add_calls[0]
        self.assertEqual(name, "Reference")
        self.assertEqual(intensities, [1.0, 2.0])
        self.assertEqual(formula, "R")
        self.assertEqual(data_type, "raman")
        self.assertEqual(peaks[0].frequency, 100.0)
        self.assertEqual(peaks[0].width_hz, 20.0)


class IdentifierErrorPropagationTest(unittest.TestCase):
    def test_match_result_accepts_legacy_matched_points_keyword(self):
        result = MatchResult("Reference", matched_points=4)

        self.assertEqual(result.compared_points, 4)
        self.assertEqual(result.matched_points, 4)

    def test_database_connection_failure_is_logged_and_raised(self):
        with (
            patch(
                "python_analyzer.core.reference_repository.sqlite3.connect",
                side_effect=sqlite3.OperationalError("database unavailable"),
            ),
            patch(
                "python_analyzer.core.reference_repository.logger.error"
            ) as log_error,
        ):
            with self.assertRaisesRegex(sqlite3.OperationalError, "database unavailable"):
                SpectrumIdentifier("/unavailable/library.db")

        log_error.assert_called_once()
        logged_args = " ".join(str(arg) for arg in log_error.call_args.args)
        self.assertIn("library.db", logged_args)
        self.assertNotIn("/unavailable", logged_args)

    def test_find_matches_reports_compared_points(self):
        identifier = SpectrumIdentifier(":memory:")
        try:
            identifier.add_reference("Reference", [1.0, 2.0, 3.0], "R")
            matches = identifier.find_matches(np.array([1.0, 2.0, 3.0, 4.0]))
        finally:
            identifier.conn.close()

        self.assertEqual(matches[0].compared_points, 3)
        self.assertEqual(matches[0].matched_points, 3)

    def test_peak_to_reference_peak_accepts_object_peaks(self):
        peak = SimpleNamespace(
            frequency=42.0,
            intensity=2.5,
            width=3.0,
            width_hz=7.5,
            area=9.0,
            snr=11.0,
        )

        reference_peak = peak_to_reference_peak(peak)

        self.assertEqual(reference_peak.frequency, 42.0)
        self.assertEqual(reference_peak.intensity, 2.5)
        self.assertEqual(reference_peak.width_hz, 7.5)

    def test_peak_to_reference_peak_rejects_non_finite_values(self):
        peak = {"frequency": float("nan"), "intensity": 1.0}

        self.assertIsNone(peak_to_reference_peak(peak))

    def test_data_type_is_normalized_to_known_values(self):
        self.assertEqual(normalize_data_type("RAMAN!"), "raman")
        self.assertEqual(normalize_data_type("unknown custom type"), "generic")

    def test_peak_matching_is_one_to_one(self):
        unknown_peaks = [
            ReferencePeak(frequency=100.0, intensity=1.0),
            ReferencePeak(frequency=101.0, intensity=1.0),
        ]
        reference_peaks = [ReferencePeak(frequency=100.5, intensity=1.0)]

        matches = find_peak_matches(
            unknown_peaks,
            reference_peaks,
            frequency_tolerance=5.0,
        )
        score = compute_peak_based_score(matches, len(unknown_peaks), len(reference_peaks))

        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0].reference_index, 0)
        self.assertLess(score, 0.95)

    def test_peak_matching_respects_tolerance(self):
        matches = find_peak_matches(
            [ReferencePeak(frequency=100.0, intensity=1.0)],
            [ReferencePeak(frequency=106.0, intensity=1.0)],
            frequency_tolerance=5.0,
        )

        self.assertEqual(matches, [])

    def test_peak_matching_log_ratio_scoring_is_symmetric(self):
        high_unknown = find_peak_matches(
            [ReferencePeak(frequency=100.0, intensity=2.0)],
            [ReferencePeak(frequency=100.0, intensity=1.0)],
            frequency_tolerance=5.0,
        )
        high_reference = find_peak_matches(
            [ReferencePeak(frequency=100.0, intensity=1.0)],
            [ReferencePeak(frequency=100.0, intensity=2.0)],
            frequency_tolerance=5.0,
        )

        self.assertEqual(high_unknown[0].score, high_reference[0].score)

    def test_peak_reference_migrates_old_database_and_round_trips(self):
        with TemporaryDirectory() as temp_dir:
            db_path = Path(temp_dir) / "library.db"
            conn = sqlite3.connect(db_path)
            try:
                conn.execute(
                    """
                    CREATE TABLE compounds (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        formula TEXT,
                        spectrum TEXT NOT NULL
                    )
                    """
                )
                conn.commit()
            finally:
                conn.close()

            identifier = SpectrumIdentifier(db_path)
            try:
                added = identifier.add_reference(
                    "PeakOnly",
                    None,
                    "P",
                    peaks=[
                        ReferencePeak(
                            frequency=100.0,
                            intensity=1.0,
                            width=2.0,
                            width_hz=20.0,
                            area=3.0,
                            snr=8.0,
                        )
                    ],
                    data_type="Raman",
                )

                row = identifier.conn.execute(
                    """
                    SELECT spectrum, peaks_json, schema_version, data_type
                    FROM compounds
                    WHERE name = ?
                    """,
                    ("PeakOnly",),
                ).fetchone()
                matches = identifier.find_peak_matches(
                    [SimpleNamespace(frequency=101.0, intensity=1.0)],
                    frequency_tolerance=5.0,
                    data_type="raman",
                )
            finally:
                identifier.conn.close()

        self.assertTrue(added)
        self.assertEqual(row[0], "[]")
        self.assertEqual(row[2], 2)
        self.assertEqual(row[3], "raman")
        self.assertEqual(json.loads(row[1])[0]["width_hz"], 20.0)
        self.assertEqual(len(matches), 1)
        self.assertEqual(matches[0].num_matched, 1)

    def test_reference_library_lists_counts_and_deletes_entries(self):
        identifier = SpectrumIdentifier(":memory:")
        try:
            identifier.add_reference("Legacy", [1.0, 2.0, 3.0], "L")
            identifier.add_reference(
                "PeakOnly",
                None,
                "P",
                peaks=[ReferencePeak(frequency=100.0, intensity=1.0)],
                data_type="raman",
            )
            identifier.conn.execute(
                """
                INSERT INTO compounds
                (name, formula, spectrum, peaks_json, schema_version, data_type)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                ("Broken", "B", "{not-json", "{not-json", 2, "unknown"),
            )
            identifier.conn.commit()

            entries = identifier.list_references()
            by_name = {entry.name: entry for entry in entries}

            self.assertEqual(by_name["Legacy"].spectrum_points, 3)
            self.assertEqual(by_name["Legacy"].peak_count, 0)
            self.assertEqual(by_name["PeakOnly"].spectrum_points, 0)
            self.assertEqual(by_name["PeakOnly"].peak_count, 1)
            self.assertEqual(by_name["PeakOnly"].data_type, "raman")
            self.assertEqual(by_name["Broken"].spectrum_points, 0)
            self.assertEqual(by_name["Broken"].peak_count, 0)
            self.assertEqual(by_name["Broken"].data_type, "generic")

            self.assertTrue(identifier.delete_reference(by_name["Legacy"].reference_id))
            self.assertFalse(identifier.delete_reference(by_name["Legacy"].reference_id))
            remaining_names = {entry.name for entry in identifier.list_references()}
        finally:
            identifier.conn.close()

        self.assertNotIn("Legacy", remaining_names)
        self.assertIn("PeakOnly", remaining_names)

    def test_peak_matching_filters_by_data_type_with_generic_fallback(self):
        identifier = SpectrumIdentifier(":memory:")
        try:
            identifier.add_reference(
                "Raman",
                None,
                "R",
                peaks=[ReferencePeak(frequency=100.0, intensity=1.0)],
                data_type="raman",
            )
            identifier.add_reference(
                "IR",
                None,
                "I",
                peaks=[ReferencePeak(frequency=100.0, intensity=1.0)],
                data_type="ir",
            )
            identifier.add_reference(
                "Generic",
                None,
                "G",
                peaks=[ReferencePeak(frequency=100.0, intensity=1.0)],
                data_type="generic",
            )

            matches = identifier.find_peak_matches(
                [ReferencePeak(frequency=100.0, intensity=1.0)],
                frequency_tolerance=5.0,
                data_type="raman",
            )
        finally:
            identifier.conn.close()

        self.assertEqual(
            {match.substance_name for match in matches},
            {"Raman", "Generic"},
        )

    def test_restore_default_delegates_repository_failure(self):
        identifier = SpectrumIdentifier.__new__(SpectrumIdentifier)
        identifier.repository = SimpleNamespace(restore_default=lambda: False)

        self.assertFalse(identifier.restore_default())


if __name__ == "__main__":
    unittest.main()
