import csv
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
import unittest

from openpyxl import load_workbook
from PIL import Image

from python_analyzer.exporters import (
    ExcelReportExporter,
    HTMLReportExporter,
    PDFMatchRow,
    PDFReportData,
    write_peaks_csv,
)
from python_analyzer.exporters.report_data import _match_to_pdf_row


def sample_report_data():
    peak = SimpleNamespace(
        frequency=120.5,
        position=42.0,
        intensity=0.84,
        width=3.2,
        width_hz=1.6,
        area=2.4,
        snr=18.0,
    )
    return PDFReportData(
        title="ChromaTsvet <Analysis> Report",
        subtitle="Spectral data and chromatogram analysis",
        summary_rows=[
            ("Source file", "sample.csv"),
            ("Data points", "128"),
            ("Peaks found", "1"),
        ],
        parameter_rows=[
            ("Sample rate", "1000 Hz"),
            ("Baseline", "improved"),
            ("Normalization", "disabled"),
        ],
        processing_passport_rows=[
            ("Generated at", "2026-07-24 12:30:00"),
            ("Analysis method", "Raman QC"),
            ("Processing warnings", "none"),
        ],
        peaks=[peak],
        matches=[
            PDFMatchRow(
                substance_name="Reference",
                formula="R",
                score="0.900",
                compared_points="1",
            )
        ],
        source_file_name="sample.csv",
        data_points_count=128,
        peaks_count=1,
    )


class ReportExportersTest(unittest.TestCase):
    def test_peak_match_diagnostics_flow_into_report_rows(self):
        row = _match_to_pdf_row(
            SimpleNamespace(
                substance_name="Reference",
                formula="R",
                method="peak",
                score=0.9123,
                compared_points=3,
                sample_coverage=0.75,
                reference_coverage=0.6,
                mean_frequency_error=0.125,
                evidence_level="strong",
            )
        )

        self.assertEqual(row.method, "peak")
        self.assertEqual(row.score, "0.912")
        self.assertEqual(row.sample_coverage, "75.0%")
        self.assertEqual(row.reference_coverage, "60.0%")
        self.assertEqual(row.mean_frequency_error, "0.1250")
        self.assertEqual(row.evidence_level, "strong")

    def test_html_and_excel_reports_write_peak_match_diagnostics(self):
        report_data = sample_report_data()
        report_data.matches = [
            PDFMatchRow(
                substance_name="Reference",
                formula="R",
                score="0.912",
                compared_points="3",
                method="peak",
                evidence_level="strong",
                sample_coverage="75.0%",
                reference_coverage="60.0%",
                mean_frequency_error="0.1250",
            )
        ]

        with TemporaryDirectory() as temp_dir:
            html_path = Path(temp_dir) / "report.html"
            excel_path = Path(temp_dir) / "report.xlsx"
            HTMLReportExporter().export(html_path, report_data)
            ExcelReportExporter().export(excel_path, report_data)
            html = html_path.read_text(encoding="utf-8")
            workbook = load_workbook(excel_path, data_only=True)

        self.assertIn("Sample coverage", html)
        self.assertIn("75.0%", html)
        self.assertIn("strong", html)
        self.assertEqual(workbook["Matches"]["C2"].value, "peak")
        self.assertEqual(workbook["Matches"]["F2"].value, "75.0%")
        self.assertEqual(workbook["Matches"]["I2"].value, "strong")

    def test_html_report_contains_escaped_report_content_and_embedded_plot(self):
        with TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            output_path = temp_path / "report.html"
            plot_path = temp_path / "plot.png"
            Image.new("RGB", (160, 90), "white").save(plot_path)

            HTMLReportExporter().export(
                output_path,
                sample_report_data(),
                plot_image_path=plot_path,
            )

            html = output_path.read_text(encoding="utf-8")

        self.assertIn("ChromaTsvet &lt;Analysis&gt; Report", html)
        self.assertIn("sample.csv", html)
        self.assertIn("Processing Passport", html)
        self.assertIn("Raman QC", html)
        self.assertIn("Reference", html)
        self.assertIn("120.5", html)
        self.assertIn("data:image/png;base64,", html)

    def test_html_report_rejects_missing_plot_without_path_leak(self):
        with TemporaryDirectory() as temp_dir:
            missing_plot = Path(temp_dir) / "missing.png"
            output_path = Path(temp_dir) / "report.html"

            with self.assertRaises(FileNotFoundError) as error_context:
                HTMLReportExporter().export(
                    output_path,
                    sample_report_data(),
                    plot_image_path=missing_plot,
                )

            self.assertNotIn(temp_dir, str(error_context.exception))
            self.assertFalse(output_path.exists())

    def test_excel_report_writes_expected_sheets_and_cells(self):
        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "report.xlsx"

            ExcelReportExporter().export(output_path, sample_report_data())

            workbook = load_workbook(output_path, data_only=True)

        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "Parameters", "Processing Passport", "Peaks", "Matches"],
        )
        self.assertEqual(workbook["Summary"]["A2"].value, "Source file")
        self.assertEqual(workbook["Summary"]["B2"].value, "sample.csv")
        self.assertEqual(workbook["Parameters"]["A2"].value, "Sample rate")
        self.assertEqual(
            workbook["Processing Passport"]["A3"].value,
            "Analysis method",
        )
        self.assertEqual(workbook["Processing Passport"]["B3"].value, "Raman QC")
        self.assertEqual(workbook["Peaks"]["A2"].value, 120.5)
        self.assertEqual(workbook["Peaks"]["G2"].value, 18.0)
        self.assertEqual(workbook["Matches"]["A2"].value, "Reference")
        self.assertEqual(workbook["Matches"]["C2"].value, "legacy cosine")
        self.assertEqual(workbook["Matches"]["D2"].value, "0.900")
        self.assertEqual(workbook["Matches"]["I2"].value, "legacy")

    def test_excel_report_neutralizes_formula_like_text(self):
        report_data = sample_report_data()
        report_data.summary_rows.append(("Operator note", '=HYPERLINK("http://x")'))
        report_data.matches = [
            PDFMatchRow(
                substance_name="=cmd",
                formula="+H2O",
                score="@score",
                compared_points="-1",
            )
        ]

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "report.xlsx"

            ExcelReportExporter().export(output_path, report_data)

            workbook = load_workbook(output_path, data_only=False)

        self.assertEqual(workbook["Summary"]["B5"].value, '\'=HYPERLINK("http://x")')
        self.assertEqual(workbook["Matches"]["A2"].value, "'=cmd")
        self.assertEqual(workbook["Matches"]["B2"].value, "'+H2O")
        self.assertEqual(workbook["Matches"]["D2"].value, "'@score")
        self.assertEqual(workbook["Matches"]["E2"].value, "'-1")

    def test_peak_csv_neutralizes_formula_like_text(self):
        peak = SimpleNamespace(
            frequency=120.5,
            position=42.0,
            intensity="=bad",
            width=3.2,
            width_hz=1.6,
            area=2.4,
            snr=18.0,
        )
        review = SimpleNamespace(status="accepted", reason="-manual note")

        with TemporaryDirectory() as temp_dir:
            output_path = Path(temp_dir) / "peaks.csv"
            write_peaks_csv(
                output_path,
                [peak],
                {
                    "source_file": '=HYPERLINK("http://x")',
                    "data_type": "@raman",
                },
                [review],
            )
            with output_path.open(newline="", encoding="utf-8") as csv_file:
                rows = list(csv.reader(csv_file))

        self.assertEqual(rows[1][0], '\'=HYPERLINK("http://x")')
        self.assertEqual(rows[1][9], "'@raman")
        self.assertEqual(rows[1][13], "'=bad")
        self.assertEqual(rows[1][19], "'-manual note")


if __name__ == "__main__":
    unittest.main()
