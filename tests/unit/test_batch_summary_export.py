from __future__ import annotations

import csv
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
import pytest

from python_analyzer.analysis.batch import (
    BATCH_STATUS_FAILED,
    BATCH_STATUS_SUCCESS,
    BatchAnalysisItem,
    BatchAnalysisSummary,
)
from python_analyzer.exporters import (
    write_batch_summary_csv,
    write_batch_summary_excel,
)


def sample_summary() -> BatchAnalysisSummary:
    return BatchAnalysisSummary(
        items=(
            BatchAnalysisItem(
                source_name="good.csv",
                status=BATCH_STATUS_SUCCESS,
                point_count=512,
                peak_count=7,
                warning_count=1,
                skipped_row_count=2,
            ),
            BatchAnalysisItem(
                source_name="bad.csv",
                status=BATCH_STATUS_FAILED,
                error_message="Unsupported or malformed spectrum data.",
            ),
        ),
        requested_count=3,
        cancelled=True,
    )


def test_batch_csv_contains_summary_and_file_rows(tmp_path):
    output_path = tmp_path / "batch.csv"

    write_batch_summary_csv(output_path, sample_summary())

    with output_path.open(encoding="utf-8-sig", newline="") as csv_file:
        rows = list(csv.reader(csv_file))

    assert rows[:3] == [
        ["metric", "value"],
        ["Requested files", "3"],
        ["Processed files", "2"],
    ]
    assert rows[5] == ["Cancelled", "yes"]
    assert rows[7][0:3] == ["source_file", "status", "point_count"]
    assert rows[8] == ["good.csv", "success", "512", "7", "1", "2", ""]
    assert rows[9][0:2] == ["bad.csv", "failed"]
    assert rows[9][6] == "Unsupported or malformed spectrum data."


def test_batch_csv_uses_utf8_bom_and_preserves_unicode(tmp_path):
    output_path = tmp_path / "unicode.csv"
    summary = BatchAnalysisSummary(
        items=(
            BatchAnalysisItem(
                source_name="спектр.csv",
                status=BATCH_STATUS_SUCCESS,
            ),
        ),
        requested_count=1,
    )

    write_batch_summary_csv(output_path, summary)

    assert output_path.read_bytes().startswith(b"\xef\xbb\xbf")
    with output_path.open(encoding="utf-8-sig", newline="") as csv_file:
        rows = list(csv.reader(csv_file))
    assert rows[8][0] == "спектр.csv"


def test_batch_excel_contains_summary_and_file_sheets(tmp_path):
    output_path = tmp_path / "batch.xlsx"

    write_batch_summary_excel(output_path, sample_summary())
    workbook = load_workbook(output_path, data_only=False)

    assert workbook.sheetnames == ["Summary", "Files"]
    assert workbook["Summary"]["B2"].value == 3
    assert workbook["Summary"]["B6"].value == "yes"
    assert workbook["Files"]["A2"].value == "good.csv"
    assert workbook["Files"]["D2"].value == 7
    assert workbook["Files"]["G3"].value == (
        "Unsupported or malformed spectrum data."
    )


@pytest.mark.parametrize("format_name", ("csv", "excel"))
def test_batch_exports_neutralize_formula_like_text(tmp_path, format_name):
    summary = BatchAnalysisSummary(
        items=(
            BatchAnalysisItem(
                source_name='=HYPERLINK("https://example.invalid")',
                status=BATCH_STATUS_FAILED,
                error_message="@SUM(1+1)",
            ),
        ),
        requested_count=1,
    )

    if format_name == "csv":
        output_path = tmp_path / "safe.csv"
        write_batch_summary_csv(output_path, summary)
        with output_path.open(encoding="utf-8-sig", newline="") as csv_file:
            rows = list(csv.reader(csv_file))
        source_value = rows[8][0]
        detail_value = rows[8][6]
    else:
        output_path = tmp_path / "safe.xlsx"
        write_batch_summary_excel(output_path, summary)
        workbook = load_workbook(output_path, data_only=False)
        source_value = workbook["Files"]["A2"].value
        detail_value = workbook["Files"]["G2"].value

    assert source_value.startswith("'=")
    assert detail_value.startswith("'@")


def test_excel_export_failure_preserves_existing_destination(tmp_path):
    output_path = tmp_path / "existing.xlsx"
    output_path.write_bytes(b"original-content")

    with (
        patch(
            "python_analyzer.exporters.batch_summary.Workbook.save",
            side_effect=OSError("simulated failure"),
        ),
        pytest.raises(OSError, match="simulated failure"),
    ):
        write_batch_summary_excel(output_path, sample_summary())

    assert output_path.read_bytes() == b"original-content"
    assert list(tmp_path.glob(".existing.xlsx.*.tmp")) == []
