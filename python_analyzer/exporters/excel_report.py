"""Excel workbook export for analysis results."""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any, Union

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from .pdf_report import PDFReportData
from .spreadsheet_safety import safe_spreadsheet_cell


PathLike = Union[str, Path]


class ExcelReportExporter:
    """Generate an XLSX workbook from prepared analysis data."""

    header_fill = PatternFill("solid", fgColor="E9EEF5")
    header_font = Font(bold=True)

    def export(self, output_path: PathLike, report_data: PDFReportData) -> None:
        """Write an Excel workbook to ``output_path``."""
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        self._write_key_value_sheet(summary_sheet, report_data.summary_rows)
        self._write_key_value_sheet(
            workbook.create_sheet("Parameters"),
            report_data.parameter_rows,
        )
        if report_data.processing_passport_rows:
            self._write_key_value_sheet(
                workbook.create_sheet("Processing Passport"),
                report_data.processing_passport_rows,
            )
        self._write_peaks_sheet(workbook.create_sheet("Peaks"), report_data.peaks)
        self._write_matches_sheet(workbook.create_sheet("Matches"), report_data.matches)

        workbook.save(Path(output_path))

    def _write_key_value_sheet(
        self,
        sheet: Worksheet,
        rows: list[tuple[str, str]],
    ) -> None:
        sheet.append(["Field", "Value"])
        for label, value in rows:
            sheet.append([str(label), self._cell_value(value)])
        self._style_header(sheet, 2)

    def _write_peaks_sheet(self, sheet: Worksheet, peaks: Any) -> None:
        sheet.append(
            [
                "frequency_hz",
                "position_bin",
                "intensity",
                "width_bins",
                "width_hz",
                "area",
                "snr",
            ]
        )
        for peak in peaks:
            sheet.append(
                [
                    self._peak_frequency(peak),
                    self._peak_value(peak, "position"),
                    self._peak_value(peak, "intensity"),
                    self._peak_value(peak, "width"),
                    self._peak_value(peak, "width_hz"),
                    self._peak_value(peak, "area"),
                    self._peak_value(peak, "snr"),
                ]
            )
        self._style_header(sheet, 7)

    def _write_matches_sheet(self, sheet: Worksheet, matches: Any) -> None:
        sheet.append(
            [
                "candidate",
                "formula",
                "method",
                "score",
                "matched_peaks",
                "sample_coverage",
                "reference_coverage",
                "mean_frequency_error",
                "evidence",
            ]
        )
        for match in matches:
            sheet.append(
                [
                    self._cell_value(match.substance_name),
                    self._cell_value(match.formula),
                    self._cell_value(match.method),
                    self._cell_value(match.score),
                    self._cell_value(match.compared_points),
                    self._cell_value(match.sample_coverage),
                    self._cell_value(match.reference_coverage),
                    self._cell_value(match.mean_frequency_error),
                    self._cell_value(match.evidence_level),
                ]
            )
        self._style_header(sheet, 9)

    def _style_header(self, sheet: Worksheet, column_count: int) -> None:
        for cell in sheet[1][:column_count]:
            cell.font = self.header_font
            cell.fill = self.header_fill

        for index in range(1, column_count + 1):
            column_letter = sheet.cell(row=1, column=index).column_letter
            sheet.column_dimensions[column_letter].width = 18

    def _peak_value(self, peak: Any, name: str) -> Any:
        return self._cell_value(getattr(peak, name, None))

    def _peak_frequency(self, peak: Any) -> Any:
        frequency = getattr(peak, "frequency", None)
        try:
            numeric_frequency = float(frequency)
        except (TypeError, ValueError):
            return self._peak_value(peak, "position")
        if not math.isfinite(numeric_frequency):
            return self._peak_value(peak, "position")
        return numeric_frequency

    def _cell_value(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (str, int, float, bool)):
            if isinstance(value, float) and not math.isfinite(value):
                return None
            return safe_spreadsheet_cell(value)
        return safe_spreadsheet_cell(str(value))
