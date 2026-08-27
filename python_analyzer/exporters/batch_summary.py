"""Safe CSV and Excel exports for compact batch-analysis summaries."""

from __future__ import annotations

from contextlib import contextmanager
import csv
import os
from pathlib import Path
import tempfile
from typing import Iterator

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from python_analyzer.analysis.batch import BatchAnalysisItem, BatchAnalysisSummary

from .spreadsheet_safety import safe_spreadsheet_cell


BATCH_FILE_HEADERS = (
    "source_file",
    "status",
    "point_count",
    "peak_count",
    "warning_count",
    "skipped_row_count",
    "details",
)


def write_batch_summary_csv(
    output_path: str | Path,
    summary: BatchAnalysisSummary,
) -> None:
    """Write a human-readable CSV summary for one batch operation."""

    with _atomic_output_path(output_path) as temporary_path:
        with temporary_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(("metric", "value"))
            for label, value in _summary_rows(summary):
                writer.writerow((label, safe_spreadsheet_cell(value)))
            writer.writerow(())
            writer.writerow(BATCH_FILE_HEADERS)
            for item in summary.items:
                writer.writerow(_safe_item_row(item))


def write_batch_summary_excel(
    output_path: str | Path,
    summary: BatchAnalysisSummary,
) -> None:
    """Write a two-sheet XLSX workbook for one batch operation."""

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(("Metric", "Value"))
    for row in _summary_rows(summary):
        summary_sheet.append(row)
    _style_sheet(summary_sheet, column_count=2)

    files_sheet = workbook.create_sheet("Files")
    files_sheet.append(BATCH_FILE_HEADERS)
    for item in summary.items:
        files_sheet.append(_safe_item_row(item))
    _style_sheet(files_sheet, column_count=len(BATCH_FILE_HEADERS))

    with _atomic_output_path(output_path) as temporary_path:
        workbook.save(temporary_path)


def _summary_rows(summary: BatchAnalysisSummary) -> tuple[tuple[str, object], ...]:
    return (
        ("Requested files", summary.requested_count),
        ("Processed files", len(summary.items)),
        ("Successful files", summary.successful_count),
        ("Failed files", summary.failed_count),
        ("Cancelled", "yes" if summary.cancelled else "no"),
    )


def _safe_item_row(item: BatchAnalysisItem) -> tuple[object, ...]:
    return tuple(
        safe_spreadsheet_cell(value)
        for value in (
            item.source_name,
            item.status,
            item.point_count,
            item.peak_count,
            item.warning_count,
            item.skipped_row_count,
            item.error_message,
        )
    )


def _style_sheet(sheet: Worksheet, *, column_count: int) -> None:
    header_fill = PatternFill("solid", fgColor="E9EEF5")
    header_font = Font(bold=True)
    for cell in sheet[1][:column_count]:
        cell.font = header_font
        cell.fill = header_fill

    for index in range(1, column_count + 1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = 20
    if column_count >= len(BATCH_FILE_HEADERS):
        sheet.column_dimensions["A"].width = 32
        sheet.column_dimensions["G"].width = 42
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


@contextmanager
def _atomic_output_path(output_path: str | Path) -> Iterator[Path]:
    """Yield a neighboring temporary path and replace the target on success."""

    destination = Path(output_path).expanduser()
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{destination.name}.",
        suffix=".tmp",
        dir=destination.parent,
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        yield temporary_path
        os.replace(temporary_path, destination)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise
